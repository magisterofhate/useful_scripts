# helpers.py
import os
import requests
import pandas as pd
from datetime import datetime, date, timedelta, timezone

from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    BASE_URL,
    API_TOKEN,
    ISSUE_QUERY,
    BASE_FILE_NAME,
)

HEADERS_YT = {
    "Authorization": f"Bearer {API_TOKEN}",
    "Accept": "application/json",
}

WEEKDAY_RU_SHORT = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]


def format_date_ru(d: date) -> str:
    """Форматируем дату для заголовка колонки: 'Пн 03.11'."""
    wd = WEEKDAY_RU_SHORT[d.weekday()]
    return f"{wd} {d:%d.%m}"


# ======================== HUB / USERS ========================================

def get_group_users_by_id(group_id: str):
    """
    Получаем логины пользователей из группы Hub по ID.
    """
    hub_base = BASE_URL.rstrip("/") + "/hub"
    url = f"{hub_base}/api/rest/usergroups/{group_id}"
    params = {"fields": "id,name,users(login,name)"}
    resp = requests.get(url, headers=HEADERS_YT, params=params)
    resp.raise_for_status()
    data = resp.json()
    users = data.get("users", [])
    logins = [u["login"] for u in users if "login" in u]
    print(f"Группа по ID '{data.get('name')}', пользователей: {len(logins)}")
    return logins


def get_group_users_by_name(group_name: str):
    """
    Получаем логины пользователей по имени группы в Hub.
    Если несколько групп — берём точное совпадение по имени, иначе первую.
    Hub возвращает объект вида:
      { "usergroups": [ {..}, {..} ], "top": 100, "skip": 0 }
    """
    hub_base = BASE_URL.rstrip("/") + "/hub"
    url = f"{hub_base}/api/rest/usergroups"
    params = {
        "query": group_name,
        "fields": "id,name,users(login,name)",
    }
    resp = requests.get(url, headers=HEADERS_YT, params=params)
    resp.raise_for_status()

    data = resp.json()

    # Hub обычно возвращает объект с ключом "usergroups"
    if isinstance(data, dict) and "usergroups" in data:
        groups = data.get("usergroups", [])
    elif isinstance(data, list):
        # На случай старого/нестандартного формата
        groups = data
    else:
        groups = []

    if not groups:
        raise RuntimeError(f"Группа '{group_name}' не найдена в Hub или нет прав на чтение групп.")

    # Ищем точное совпадение по имени, иначе берём первую
    group = next((g for g in groups if g.get("name") == group_name), groups[0])

    users = group.get("users", []) or []
    logins = [u["login"] for u in users if isinstance(u, dict) and u.get("login")]
    print(f"Группа по имени '{group.get('name')}', пользователей: {len(logins)}")
    return logins


def fetch_users_map():
    """
    Получаем карту login -> fullName (ФИО).
    Если fullName пустой, будем использовать login.
    """
    users_map = {}
    url = BASE_URL.rstrip("/") + "/api/users"
    skip = 0
    while True:
        params = {
            "fields": "login,fullName",
            "$top": 100,
            "$skip": skip,
        }
        resp = requests.get(url, headers=HEADERS_YT, params=params)
        resp.raise_for_status()
        batch = resp.json()
        if not batch:
            break
        for u in batch:
            login = u.get("login")
            full_name = (u.get("fullName") or "").strip()
            if login:
                users_map[login] = full_name or login
        if len(batch) < 100:
            break
        skip += len(batch)

    print(f"Загружено пользователей (login->ФИО): {len(users_map)}")
    return users_map


# ======================== WORK ITEMS =========================================

def fetch_work_items_for_users(user_logins, start_date: date, end_date: date, issue_query: str = ""):
    """
    Тянем work items по всем пользователям за период start_date..end_date,
    с фильтром по задачам issue_query (если задан).
    """
    all_items = []
    query = issue_query or ISSUE_QUERY or ""

    for login in user_logins:
        skip = 0
        while True:
            params = {
                "fields": (
                    "author(login,fullName),date,"
                    "duration(minutes),issue(idReadable,summary)"
                ),
                "author": login,
                "startDate": start_date.isoformat(),
                "endDate": end_date.isoformat(),
                "query": query,
                "$top": 100,
                "$skip": skip,
            }
            url = BASE_URL.rstrip("/") + "/api/workItems"
            resp = requests.get(url, headers=HEADERS_YT, params=params)
            resp.raise_for_status()
            batch = resp.json()
            if not batch:
                break

            all_items.extend(batch)

            if len(batch) < 100:
                break
            skip += len(batch)

        print(f"{login}: загружено work items: {skip}+")
    return all_items


# ======================== MATRIX / DATAFRAME ================================

def build_timesheet_matrix(
    work_items,
    user_logins,
    users_map,
    start_date: date,
    end_date: date,
):
    """
    Из списка work items строим pandas-DataFrame вида:
        ФИО × дата → часы (2 знака после запятой)
    - по выходным нули НЕ ставим (ячейки пустые),
    - индекс — ФИО (fullName),
    - добавляем колонку 'Итого'.
    """
    records = []
    for wi in work_items:
        author = wi.get("author") or {}
        login = author.get("login", "unknown")
        full_name_api = (author.get("fullName") or "").strip()
        display_name = full_name_api or users_map.get(login, login)

        duration = (wi.get("duration") or {}).get("minutes", 0)

        ts = wi.get("date")
        if ts is None:
            continue
        d = datetime.fromtimestamp(ts / 1000, tz=timezone.utc).date()

        records.append({
            "user": display_name,
            "login": login,
            "date": d,
            "minutes": duration,
        })

    # полный список дат периода
    all_days = []
    cur = start_date
    while cur <= end_date:
        all_days.append(cur)
        cur += timedelta(days=1)

    # целевые имена (ФИО) в порядке логинов
    target_names = [users_map.get(login, login) for login in user_logins]

    # Если нет записей вовсе — пустая матрица с нулями (в выходные пусто)
    if not records:
        print("Нет списаний за указанный период.")
        df = pd.DataFrame(0.0, index=target_names, columns=all_days)
        df.index.name = "ФИО"
        # по выходным — пустые ячейки
        for col in df.columns:
            if isinstance(col, date) and col.weekday() >= 5:
                df[col] = pd.NA
        df["Итого"] = 0.0
        return df.round(2)

    df = pd.DataFrame(records)

    # pivot: минуты → часы
    pivot = df.pivot_table(
        index="user",
        columns="date",
        values="minutes",
        aggfunc="sum",
        fill_value=0
    ) / 60.0  # часы

    # добавляем недостающие даты
    for d in all_days:
        if d not in pivot.columns:
            pivot[d] = 0.0

    # добавляем всех пользователей
    for name in target_names:
        if name not in pivot.index:
            pivot.loc[name] = 0.0

    # сортировка: даты и строки в нужном порядке
    date_cols = sorted([c for c in pivot.columns if isinstance(c, date)])
    pivot = pivot[date_cols]
    pivot = pivot.reindex(target_names)

    # по выходным вместо 0 делаем пустые ячейки
    for col in pivot.columns:
        if isinstance(col, date) and col.weekday() >= 5:
            pivot.loc[pivot[col] == 0, col] = pd.NA

    # добавляем итог по строкам
    pivot["Итого"] = pivot.sum(axis=1, numeric_only=True)

    # округляем до двух знаков
    pivot = pivot.round(2)
    pivot.index.name = "ФИО"
    return pivot


def build_details_sheet(work_items, users_map):
    """
    Формирует подробный лист:
    ФИО / login / дата / часы / issue / summary
    """
    detail_records = []
    for wi in work_items:
        author = wi.get("author") or {}
        login = author.get("login", "")
        full_name_api = (author.get("fullName") or "").strip()
        display_name = full_name_api or users_map.get(login, login)

        ts = wi.get("date")
        if ts is None:
            continue
        d = datetime.fromtimestamp(ts / 1000, tz=timezone.utc).date()
        minutes = (wi.get("duration") or {}).get("minutes", 0)

        issue = wi.get("issue") or {}

        detail_records.append({
            "ФИО": display_name,
            "login": login,
            "date": d,
            "hours": round(minutes / 60.0, 2),
            "issue": issue.get("idReadable", ""),
            "summary": issue.get("summary", ""),
        })

    if not detail_records:
        return pd.DataFrame(
            columns=["ФИО", "login", "date", "hours", "issue", "summary"]
        )
    return pd.DataFrame(detail_records)


# ======================== EXCEL / FORMATTING ================================

def build_output_filename(start_date: date, end_date: date) -> str:
    """
    Формирует базовое имя файла: timesheet_YYYY-MM-DD_YYYY-MM-DD.xlsx
    """
    return f"{BASE_FILE_NAME}_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"


def get_available_filename(base_name: str) -> str:
    """
    Проверяет, существует ли файл. Если существует — добавляет (1), (2), ...
    Возвращает доступное имя файла.
    """
    if not os.path.exists(base_name):
        return base_name

    name, ext = os.path.splitext(base_name)

    i = 1
    while True:
        new_name = f"{name} ({i}){ext}"
        if not os.path.exists(new_name):
            return new_name
        i += 1


def write_excel_with_formatting(timesheet_df, details_df, start_date: date, end_date: date):
    """
    Пишем Excel:
    - имя файла: BASE_FILE_NAME_YYYY-MM-DD_YYYY-MM-DD.xlsx (+ (1), (2), ... при коллизии);
    - лист Timesheet:
        - колонка A = ФИО, ширина под самое длинное ФИО;
        - колонка B = Группа (мы добавили её в main.py);
        - выходные дни подсвечены светло-оранжевым;
        - нули в будни подсвечены бледно-красным;
        - в выходные дни нули не показываются (ячейки пустые);
        - между разными группами рисуется жирная горизонтальная линия;
    - лист Details (подробные списания);
    - лист group_members:
        - ФИО, Группа, Списано часов за период, Часов в периоде;
        - колонка ФИО по ширине самого длинного ФИО;
        - заголовки колонок с временем расширены;
        - Списано > часов в периоде → зелёным;
        - Списано < 80% часов в периоде → #F08080.
    """
    base_filename = build_output_filename(start_date, end_date)
    filename = get_available_filename(base_filename)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # Основная таблица
        timesheet_df.to_excel(writer, sheet_name="Timesheet")
        # Детализация
        details_df.to_excel(writer, sheet_name="Details", index=False)

        wb = writer.book
        ws = writer.sheets["Timesheet"]

        n_rows, n_cols = timesheet_df.shape  # строки = сотрудники, колонки = Группа + даты + "Итого"

        header_row = 1         # строка с заголовками
        data_start_row = 2     # первая строка данных
        first_data_col = 2     # первая колонка данных (столбец A = индекс ФИО, B = "Группа")

        # Цвета
        weekend_fill = PatternFill(
            start_color="FFF4CC", end_color="FFF4CC", fill_type="solid"
        )
        zero_fill = PatternFill(
            start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"
        )

        # 0. Ширина колонки ФИО (A) = длина макс ФИО + немного воздуха
        max_name_len = max((len(str(name)) for name in timesheet_df.index), default=10)
        ws.column_dimensions["A"].width = max_name_len + 2

        # 1. Локализуем заголовки дат и подсвечиваем выходные столбцы
        for col_idx, col_name in enumerate(timesheet_df.columns):
            excel_col = first_data_col + col_idx
            col_letter = get_column_letter(excel_col)
            cell = ws[f"{col_letter}{header_row}"]

            if isinstance(col_name, date):
                # локализованный заголовок
                cell.value = format_date_ru(col_name)

                # если выходной — красим весь столбец
                if col_name.weekday() >= 5:
                    cell.fill = weekend_fill
                    for row in range(data_start_row, data_start_row + n_rows):
                        ws.cell(row=row, column=excel_col).fill = weekend_fill
            else:
                # "Группа", "Итого" — оставляем как есть
                pass

        # 2. Подсветка нулей бледно-красным (кроме выходных столбцов)
        for row_idx in range(n_rows):
            excel_row = data_start_row + row_idx

            for col_idx, col_name in enumerate(timesheet_df.columns):
                excel_col = first_data_col + col_idx
                cell = ws.cell(row=excel_row, column=excel_col)

                # выходной?
                is_weekend = isinstance(col_name, date) and col_name.weekday() >= 5

                # для выходных нули мы уже заменили на NaN в DataFrame,
                # так что здесь трогаем только будни и "Итого"
                if not is_weekend and isinstance(cell.value, (int, float)) and cell.value == 0:
                    cell.fill = zero_fill

        # 3. Жирная черта между группами (если групп более одной)
        if "Группа" in timesheet_df.columns:
            groups = list(timesheet_df["Группа"])
            unique_groups = {g for g in groups if g not in (None, "", " ")}
            if len(unique_groups) > 1:
                thick_side = Side(style="thick")
                for idx in range(len(groups) - 1):
                    if groups[idx] != groups[idx + 1]:
                        excel_row = data_start_row + idx
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=excel_row, column=col)
                            b = cell.border
                            cell.border = Border(
                                left=b.left,
                                right=b.right,
                                top=b.top,
                                bottom=thick_side,
                            )

        # 4. Лист group_members
        # считаем число рабочих дней в периоде (Пн–Пт)
        work_days = 0
        cur = start_date
        while cur <= end_date:
            if cur.weekday() < 5:
                work_days += 1
            cur += timedelta(days=1)
        hours_in_period = work_days * 8

        summary_rows = []
        for fio in timesheet_df.index:
            # "Итого" — общие часы по человеку
            total_hours = timesheet_df.loc[fio, "Итого"]
            try:
                total_hours_val = float(total_hours) if pd.notna(total_hours) else 0.0
            except Exception:
                total_hours_val = 0.0

            if "Группа" in timesheet_df.columns:
                group_val = timesheet_df.loc[fio, "Группа"]
            else:
                group_val = ""

            summary_rows.append({
                "ФИО": fio,
                "Группа": group_val,
                "Списано часов за период": total_hours_val,
                "Часов в периоде": hours_in_period,
            })

        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="group_members", index=False)

        ws_summary = writer.sheets["group_members"]

        # Ширина колонки ФИО на group_members
        if not summary_df.empty:
            max_name_len_summary = max((len(str(v)) for v in summary_df["ФИО"]), default=10)
            ws_summary.column_dimensions["A"].width = max_name_len_summary + 2

        # Чтобы заголовки колонок с временем были видны целиком — расширим их ширину
        # Найдём индексы нужных колонок
        col_index_map = {name: idx + 1 for idx, name in enumerate(summary_df.columns)}
        spent_col_idx = col_index_map.get("Списано часов за период")
        period_col_idx = col_index_map.get("Часов в периоде")

        # Ширину под заголовки (просто делаем достаточно большой)
        if spent_col_idx:
            col_letter = get_column_letter(spent_col_idx)
            ws_summary.column_dimensions[col_letter].width = max(
                len("Списано часов за период") + 2, 25
            )
        if period_col_idx:
            col_letter = get_column_letter(period_col_idx)
            ws_summary.column_dimensions[col_letter].width = max(
                len("Часов в периоде") + 2, 22
            )

        # Подсветка по условиям:
        # - зелёным, если списано > часов в периоде
        # - #F08080, если списано < 80% часов периода
        green_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        low_fill = PatternFill(
            start_color="F08080", end_color="F08080", fill_type="solid"
        )

        if spent_col_idx and period_col_idx:
            n_summary_rows = len(summary_df)
            for i in range(n_summary_rows):
                excel_row = 2 + i  # данные начинаются со 2-й строки

                spent_cell = ws_summary.cell(row=excel_row, column=spent_col_idx)
                period_cell = ws_summary.cell(row=excel_row, column=period_col_idx)

                spent_val = spent_cell.value
                period_val = period_cell.value

                if isinstance(spent_val, (int, float)) and isinstance(period_val, (int, float)):
                    if spent_val > period_val:
                        # переработка: зелёным
                        spent_cell.fill = green_fill
                    elif spent_val < 0.8 * period_val:
                        # менее 80%: #F08080
                        spent_cell.fill = low_fill

    print(f"Готово, Excel сохранён как: {filename}")
