# helpers.py
import requests
import pandas as pd
from datetime import datetime, date, timedelta, timezone

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from config import (
    BASE_URL,
    API_TOKEN,
    START_DATE,
    END_DATE,
    ISSUE_QUERY,
    USE_GROUP,
    GROUP_ID,
    USER_LOGINS,
    OUTPUT_XLSX,
)

HEADERS_YT = {
    "Authorization": f"Bearer {API_TOKEN}",
    "Accept": "application/json",
}


# ======================== HUB / USERS ========================================

def get_group_users_from_hub():
    """
    Получаем логины пользователей из группы Hub.
    Для YouTrack Cloud hub-URL обычно: BASE_URL + '/hub'
    """
    hub_base = BASE_URL.rstrip("/") + "/hub"
    url = f"{hub_base}/api/rest/usergroups/{GROUP_ID}"
    params = {"fields": "id,name,users(login,name)"}
    resp = requests.get(url, headers=HEADERS_YT, params=params)
    resp.raise_for_status()
    data = resp.json()
    users = data.get("users", [])
    logins = [u["login"] for u in users if "login" in u]
    print(f"Группа '{data.get('name')}', пользователей: {len(logins)}")
    return logins


def fetch_users_map():
    """
    Получаем карту login -> fullName (ФИО).
    Если fullName пустой, будем использовать login.
    """
    users_map = {}
    url = BASE_URL.rstrip("/") + "/api/users"
    skip = 0
    while True:
        params = {
            "fields": "login,fullName",
            "$top": 100,
            "$skip": skip,
        }
        resp = requests.get(url, headers=HEADERS_YT, params=params)
        resp.raise_for_status()
        batch = resp.json()
        if not batch:
            break
        for u in batch:
            login = u.get("login")
            full_name = (u.get("fullName") or "").strip()
            if login:
                users_map[login] = full_name or login
        if len(batch) < 100:
            break
        skip += len(batch)

    print(f"Загружено пользователей (login->FIO): {len(users_map)}")
    return users_map


# ======================== WORK ITEMS =========================================

def fetch_work_items_for_users(user_logins):
    """
    Тянем work items по всем пользователям за период START_DATE..END_DATE,
    с фильтром по задачам ISSUE_QUERY.
    """
    all_items = []
    for login in user_logins:
        skip = 0
        while True:
            params = {
                "fields": (
                    "author(login,fullName),date,"
                    "duration(minutes),issue(idReadable,summary)"
                ),
                "author": login,
                "startDate": START_DATE,
                "endDate": END_DATE,
                "query": ISSUE_QUERY,
                "$top": 100,
                "$skip": skip,
            }
            url = BASE_URL.rstrip("/") + "/api/workItems"
            resp = requests.get(url, headers=HEADERS_YT, params=params)
            resp.raise_for_status()
            batch = resp.json()
            if not batch:
                break

            all_items.extend(batch)

            if len(batch) < 100:
                break
            skip += len(batch)

        print(f"{login}: загружено work items: {skip}+")
    return all_items


# ======================== MATRIX / DATAFRAME ================================

def build_timesheet_matrix(work_items, user_logins, users_map):
    """
    Из списка work items строим pandas-DataFrame вида:
        ФИО × дата → часы (2 знака после запятой)
    - по выходным нули НЕ ставим (ячейки будут пустыми),
    - к выходным применим особую заливку в Excel,
    - индекс — ФИО (fullName).
    """
    records = []
    for wi in work_items:
        author = wi.get("author") or {}
        login = author.get("login", "unknown")
        full_name_api = (author.get("fullName") or "").strip()
        display_name = full_name_api or users_map.get(login, login)

        duration = (wi.get("duration") or {}).get("minutes", 0)

        ts = wi.get("date")
        if ts is None:
            continue
        d = datetime.fromtimestamp(ts / 1000, tz=timezone.utc).date()

        records.append({
            "user": display_name,
            "login": login,
            "date": d,
            "minutes": duration,
        })

    # Если нет записей вообще — делаем пустую матрицу (все нули/пусто)
    if not records:
        print("Нет списаний за указанный период.")
        start = datetime.strptime(START_DATE, "%Y-%m-%d").date()
        end = datetime.strptime(END_DATE, "%Y-%m-%d").date()
        days = []
        cur = start
        while cur <= end:
            days.append(cur)
            cur += timedelta(days=1)

        # индекс — ФИО (по карте users_map, fallback login)
        index_names = [users_map.get(l, l) for l in user_logins]

        df = pd.DataFrame(0.0, index=index_names, columns=days)
        df.index.name = "ФИО"

        # по выходным сделаем пустые ячейки вместо нулей
        for col in df.columns:
            if isinstance(col, date) and col.weekday() >= 5:
                df[col] = pd.NA

        df["Итого"] = 0.0
        return df

    df = pd.DataFrame(records)

    # pivot: минуты → часы, пока без Total
    pivot = df.pivot_table(
        index="user",
        columns="date",
        values="minutes",
        aggfunc="sum",
        fill_value=0
    ) / 60.0  # часы

    # период целиком
    start = datetime.strptime(START_DATE, "%Y-%m-%d").date()
    end = datetime.strptime(END_DATE, "%Y-%m-%d").date()
    all_days = []
    cur = start
    while cur <= end:
        all_days.append(cur)
        cur += timedelta(days=1)

    # добавляем недостающие даты
    for d in all_days:
        if d not in pivot.columns:
            pivot[d] = 0.0

    # индекс должен содержать всех нужных пользователей,
    # даже если у них нет ни одного списания в этот период.
    target_names = [users_map.get(login, login) for login in user_logins]
    for name in target_names:
        if name not in pivot.index:
            pivot.loc[name] = 0.0

    # сортировка дат по возрастанию
    date_cols = sorted([c for c in pivot.columns if isinstance(c, date)])
    pivot = pivot[date_cols]

    # порядок строк = как в списке user_logins
    pivot = pivot.reindex(target_names)

    # по выходным вместо 0 делаем пустые ячейки
    for col in pivot.columns:
        if isinstance(col, date) and col.weekday() >= 5:
            pivot.loc[pivot[col] == 0, col] = pd.NA

    # добавляем итог по человеку (учитывая, что NaN не считается)
    pivot["Итого"] = pivot.sum(axis=1, numeric_only=True)

    # округляем до двух знаков
    pivot = pivot.round(2)

    pivot.index.name = "ФИО"
    return pivot


def build_details_sheet(work_items, users_map):
    """
    Формирует подробный лист:
    ФИО / login / дата / часы / issue / summary
    """
    detail_records = []
    for wi in work_items:
        author = wi.get("author") or {}
        login = author.get("login", "")
        full_name_api = (author.get("fullName") or "").strip()
        display_name = full_name_api or users_map.get(login, login)

        ts = wi.get("date")
        if ts is None:
            continue
        d = datetime.fromtimestamp(ts / 1000, tz=timezone.utc).date()
        minutes = (wi.get("duration") or {}).get("minutes", 0)

        issue = wi.get("issue") or {}

        detail_records.append({
            "ФИО": display_name,
            "login": login,
            "date": d,
            "hours": round(minutes / 60.0, 2),
            "issue": issue.get("idReadable", ""),
            "summary": issue.get("summary", ""),
        })

    if not detail_records:
        return pd.DataFrame(
            columns=["ФИО", "login", "date", "hours", "issue", "summary"]
        )
    return pd.DataFrame(detail_records)


# ======================== EXCEL / FORMATTING ================================

def write_excel_with_formatting(timesheet_df, details_df):
    """
    Пишем Excel:
    - лист 'Timesheet' с заливкой выходных колонок (светло-оранжевый),
    - нулевые значения (0) подсвечиваем бледно-красным,
      при этом для выходных нули уже превращены в пустые ячейки (NaN),
    - лист 'Details' без особого форматирования.
    """
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        # Основная таблица
        timesheet_df.to_excel(writer, sheet_name="Timesheet")
        # Детализация
        details_df.to_excel(writer, sheet_name="Details", index=False)

        wb = writer.book
        ws = writer.sheets["Timesheet"]

        n_rows, n_cols = timesheet_df.shape  # n_rows = кол-во людей, n_cols = дней + 'Итого'

        header_row = 1         # строка заголовков DataFrame
        data_start_row = 2     # первая строка данных
        index_col = 1          # колонка с ФИО
        first_data_col = 2     # первая дата (или первый столбец значений)

        # Заливка для выходных (светло-оранжевый)
        weekend_fill = PatternFill(
            start_color="FFF4CC", end_color="FFF4CC", fill_type="solid"
        )
        # Заливка для нулей (бледно-красная)
        zero_fill = PatternFill(
            start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"
        )

        # 1. Подсвечиваем выходные столбцы (по заголовкам-датам)
        for col_idx, col_name in enumerate(timesheet_df.columns):
            # col_idx: 0..n_cols-1, Excel-колонка = first_data_col + col_idx
            if isinstance(col_name, date) and col_name.weekday() >= 5:
                excel_col = first_data_col + col_idx
                col_letter = get_column_letter(excel_col)

                # Заголовок
                ws[f"{col_letter}{header_row}"].fill = weekend_fill

                # Все ячейки в этом столбце
                for row in range(data_start_row, data_start_row + n_rows):
                    ws.cell(row=row, column=excel_col).fill = weekend_fill

        # 2. Подсвечиваем нули бледно-красным (кроме выходных столбцов)
        for row_idx in range(n_rows):
            excel_row = data_start_row + row_idx

            for col_idx, col_name in enumerate(timesheet_df.columns):
                excel_col = first_data_col + col_idx

                # Пропускаем выходные столбцы (там нули уже превратили в NaN)
                if isinstance(col_name, date) and col_name.weekday() >= 5:
                    continue

                cell = ws.cell(row=excel_row, column=excel_col)
                # если значение 0 (int/float), подсвечиваем
                if isinstance(cell.value, (int, float)) and cell.value == 0:
                    cell.fill = zero_fill

    print(f"Готово, Excel сохранён как {OUTPUT_XLSX}")
