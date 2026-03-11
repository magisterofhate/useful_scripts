from __future__ import annotations

import os
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from datetime import datetime, date
from typing import Optional, List, Tuple
from bisect import bisect_right


def get_unique_filename(path: str) -> str:
    """
    Windows-like: file.xlsx, file (1).xlsx, file (2).xlsx
    """
    if not os.path.exists(path):
        return path

    directory = os.path.dirname(path)
    filename = os.path.basename(path)
    base, ext = os.path.splitext(filename)

    # if base already ends with (n), continue from n+1
    m = re.match(r"^(.*)\((\d+)\)$", base.strip())
    if m:
        base_name = m.group(1).strip()
        counter = int(m.group(2)) + 1
    else:
        base_name = base
        counter = 1

    while True:
        new_filename = f"{base_name} ({counter}){ext}"
        new_path = os.path.join(directory, new_filename)
        if not os.path.exists(new_path):
            return new_path
        counter += 1


def write_versions_sheet(wb, versions: List[Tuple[str, str]], sheet_name: str = "Versions") -> None:
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws: Worksheet = wb.create_sheet(title=sheet_name)
    ws["A1"] = "Version"
    ws["B1"] = "Release date"

    for i, (ver, rel_date) in enumerate(versions, start=2):
        ws.cell(row=i, column=1, value=ver)
        ws.cell(row=i, column=2, value=rel_date)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14


def highlight_missing_ps_version(
    wb,
    sheet_name: str,
    ps_links_col_name: str,
    ps_version_col_name: str,
) -> int:
    """
    If PS links exist but PS_Версия is empty -> coral fill.
    Returns count of colored cells.
    """
    ws = wb[sheet_name]

    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    ps_links_col = header.get(ps_links_col_name)
    ps_ver_col = header.get(ps_version_col_name)
    if ps_links_col is None or ps_ver_col is None:
        raise RuntimeError(f"Не найдены колонки '{ps_links_col_name}' и/или 'PS_{ps_version_col_name}'")

    coral_fill = PatternFill(fill_type="solid", start_color="FF7F50", end_color="FF7F50")
    coral_count = 0

    for r in range(2, ws.max_row + 1):
        links_val = ws.cell(row=r, column=ps_links_col).value
        ver_cell = ws.cell(row=r, column=ps_ver_col)
        ver_val = ver_cell.value

        has_links = bool(str(links_val).strip()) if links_val is not None else False
        has_ver = bool(str(ver_val).strip()) if ver_val is not None else False

        if has_links and not has_ver:
            ver_cell.fill = coral_fill
            coral_count += 1

    return coral_count


def _find_col(ws, name: str) -> int:
    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    if name not in header:
        raise RuntimeError(f"Не найдена колонка '{name}' на листе '{ws.title}'")
    return header[name]


def _parse_ymd(v) -> Optional[date]:
    """
    Accepts:
      - 'YYYY-MM-DD' string
      - datetime/date (if Excel stored as date)
    """
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


def _cell_str(ws, row: int, col: int) -> str:
    value = ws.cell(row=row, column=col).value
    return str(value).strip() if value is not None else ""


def _is_cancelled_status(status: str) -> bool:
    return status.strip().lower().startswith("аннули")


def fill_fix_version_from_versions(
    wb,
    *,
    defects_sheet_name: str = "Defects",
    versions_sheet_name: str = "Versions",
    status_col_name: str = "Статус",
    resolved_col_name: str = "Resolved",
    release_col_name: str = "Release",
    fix_version_col_name: str = "Fix version",
    versions_version_col_name: str = "Version",
    versions_date_col_name: str = "Release date",
) -> int:
    """
    Логика заполнения Fix version:

    1. Если статус = Аннулирована -> ничего не делаем
    2. Если Fix version уже заполнен -> ничего не делаем
    3. Если нет Resolved -> ничего не делаем
    4. Если есть Release -> Fix version = Release
    5. Иначе ищем первую версию из листа Versions, у которой
       Release date > Resolved, и ставим её в Fix version

    Возвращает количество заполненных ячеек Fix version.
    """
    if defects_sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Нет листа '{defects_sheet_name}'")
    if versions_sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Нет листа '{versions_sheet_name}'")

    ws_def = wb[defects_sheet_name]
    ws_ver = wb[versions_sheet_name]

    # Колонки Defects
    col_status = _find_col(ws_def, status_col_name)
    col_resolved = _find_col(ws_def, resolved_col_name)
    col_release = _find_col(ws_def, release_col_name)
    col_fix = _find_col(ws_def, fix_version_col_name)

    # Колонки Versions
    col_version = _find_col(ws_ver, versions_version_col_name)
    col_release_date = _find_col(ws_ver, versions_date_col_name)

    # Читаем версии и сортируем по дате релиза
    versions = []
    for r in range(2, ws_ver.max_row + 1):
        version_name = _cell_str(ws_ver, r, col_version)
        release_dt = _parse_ymd(ws_ver.cell(row=r, column=col_release_date).value)
        if version_name and release_dt:
            versions.append((release_dt, version_name))

    versions.sort(key=lambda x: x[0])

    if not versions:
        return 0

    dates = [d for d, _ in versions]
    names = [n for _, n in versions]

    filled = 0

    for r in range(2, ws_def.max_row + 1):
        status_str = _cell_str(ws_def, r, col_status)
        if _is_cancelled_status(status_str):
            continue

        fix_cell = ws_def.cell(row=r, column=col_fix)
        current_fix = str(fix_cell.value).strip() if fix_cell.value is not None else ""
        if current_fix:
            continue

        resolved_dt = _parse_ymd(ws_def.cell(row=r, column=col_resolved).value)
        if resolved_dt is None:
            continue

        release_str = _cell_str(ws_def, r, col_release)
        if release_str:
            fix_cell.value = release_str
            filled += 1
            continue

        idx = bisect_right(dates, resolved_dt)
        if idx < len(dates):
            fix_cell.value = names[idx]
            filled += 1

    return filled


def fill_affected_version_from_versions(
    wb,
    *,
    defects_sheet_name: str = "Defects",
    versions_sheet_name: str = "Versions",
    created_col_name: str = "Created",
    affected_version_col_name: str = "Affected version",
    versions_version_col_name: str = "Version",
    versions_date_col_name: str = "Release date",
) -> int:
    """
    Логика заполнения Affected version:

    1. Если Affected version уже заполнен (например, из PS_Версия) -> ничего не делаем
    2. Если нет Created -> ничего не делаем
    3. Иначе ищем первую версию из листа Versions, у которой
       Release date > Created, и ставим её в Affected version

    Возвращает количество заполненных ячеек Affected version.
    """
    if defects_sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Нет листа '{defects_sheet_name}'")
    if versions_sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Нет листа '{versions_sheet_name}'")

    ws_def = wb[defects_sheet_name]
    ws_ver = wb[versions_sheet_name]

    # Колонки Defects
    col_created = _find_col(ws_def, created_col_name)
    col_affected = _find_col(ws_def, affected_version_col_name)

    # Колонки Versions
    col_version = _find_col(ws_ver, versions_version_col_name)
    col_release_date = _find_col(ws_ver, versions_date_col_name)

    # Читаем версии и сортируем по дате релиза
    versions = []
    for r in range(2, ws_ver.max_row + 1):
        version_name = _cell_str(ws_ver, r, col_version)
        release_dt = _parse_ymd(ws_ver.cell(row=r, column=col_release_date).value)
        if version_name and release_dt:
            versions.append((release_dt, version_name))

    versions.sort(key=lambda x: x[0])

    if not versions:
        return 0

    dates = [d for d, _ in versions]
    names = [n for _, n in versions]

    filled = 0

    for r in range(2, ws_def.max_row + 1):
        affected_cell = ws_def.cell(row=r, column=col_affected)
        current_affected = str(affected_cell.value).strip() if affected_cell.value is not None else ""
        if current_affected:
            continue

        created_dt = _parse_ymd(ws_def.cell(row=r, column=col_created).value)
        if created_dt is None:
            continue

        idx = bisect_right(dates, created_dt)
        if idx < len(dates):
            affected_cell.value = names[idx]
            filled += 1

    return filled


def export_excel(
    df: pd.DataFrame,
    out_path: str,
    *,
    versions: Optional[List[Tuple[str, str]]] = None,
    main_sheet_name: str = "Defects",
    versions_sheet_name: str = "Versions",
    ps_links_col_name: str = "PS links (IDs)",
    ps_version_col_name: str = "PS_Версия",
) -> tuple[str, int, int, int]:
    """
    Writes df -> xlsx, adds Versions sheet, highlights missing PS versions.
    Returns (final_path, coral_count).
    """
    final_path = get_unique_filename(out_path)
    df.to_excel(final_path, index=False, sheet_name=main_sheet_name)

    wb = load_workbook(final_path)

    # Add versions sheet if provided
    if versions is not None:
        write_versions_sheet(wb, versions, sheet_name=versions_sheet_name)

    affected_filled = 0
    if versions is not None:
        affected_filled = fill_affected_version_from_versions(
            wb,
            defects_sheet_name=main_sheet_name,
            versions_sheet_name=versions_sheet_name,
            created_col_name="Created",
            affected_version_col_name="Affected version",
            versions_version_col_name="Version",
            versions_date_col_name="Release date",
        )

    # Автозаполнение Fix version
    fix_filled = 0
    if versions is not None:
        fix_filled = fill_fix_version_from_versions(
            wb,
            defects_sheet_name=main_sheet_name,
            versions_sheet_name=versions_sheet_name,
            resolved_col_name="Resolved",
            status_col_name="Status",
            release_col_name="Release",
            fix_version_col_name="Fix version",
            versions_version_col_name="Version",
            versions_date_col_name="Release date",
        )

    coral_count = highlight_missing_ps_version(
        wb, sheet_name=main_sheet_name,
        ps_links_col_name=ps_links_col_name,
        ps_version_col_name=ps_version_col_name,
    )

    wb.save(final_path)
    return final_path, coral_count, fix_filled, affected_filled
