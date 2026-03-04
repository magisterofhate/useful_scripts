import requests
import os
import re
import argparse
import sys
import pandas as pd
from datetime import datetime, timezone, date
from typing import Any, Dict, List, Optional
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =======================
# CONFIG
# =======================
BASE_URL = "https://youtrack.ispsystem.net/"  # без /api
TOKEN = "perm-YS5taWxpbmV2c2tpaQ==.NTgtOTU=.vkjYV9lHy4hFn2HrNvXfzAtSSNUbSM"  # персональный токен

# Пагинация
PAGE_SIZE = 200       # сколько задач за страницу (max обычно 100–200)
MAX_PAGES = 2      # например 5; если None — выгружать всё

# project = "VM"
DEFECT_TYPE = "Ошибка"  # или "Bug" / как у вас в Type
PS_PROJECT = "PS"

# Названия полей в YouTrack (как в интерфейсе)
FIELD_STATUS = "State"  # обычно State/Статус
FIELD_PRIORITY = "Priority"
FIELD_RELEASE = "Релиз"

# Поле версии в проекте PS (как в интерфейсе)
PS_VERSION_FIELD = "Версия"  # если у вас называется иначе — поправь

# Диапазон (опционально) — можно оставить None
CREATED_FROM = "2023-01-01"  # "2024-01-01"
CREATED_TO = None  # "2025-12-31"

RESOLVED_CUTOFF = "2024-07-01"  # всё что resolved раньше — исключаем

# OUT_XLSX = "vm_defects_with_ps_links.xlsx"

# =======================
# Helpers
# =======================
HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Accept": "application/json",
}


def yt_dt(ms: Optional[int]) -> Optional[str]:
    """Convert YouTrack timestamp (ms) to ISO date string YYYY-MM-DD."""
    if ms is None:
        return None
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).date().isoformat()


def calc_lifetime(created_str, resolved_str):
    """
    Calculate working days between Created and Resolved.
    If resolved is empty -> return empty string.
    """
    if not created_str or not resolved_str:
        return ""

    created = np.datetime64(created_str)
    resolved = np.datetime64(resolved_str)

    return int(np.busday_count(created, resolved))


def calc_quarter_month(created_str: str):
    """
    From Created (YYYY-MM-DD) returns:
      Quarter: 'Qx YYYY'
      Month:   'Mon YYYY' (e.g., 'Sep 2024')
    If created_str is empty -> ('', '')
    """
    if not created_str:
        return "", ""

    dt = datetime.strptime(created_str, "%Y-%m-%d").date()
    q = (dt.month - 1) // 3 + 1
    quarter = f"Q{q} {dt.year}"
    month = dt.strftime("%b %Y")  # Jan/Feb/Mar... in English locale
    return quarter, month

def normalize_ps_version(v: str) -> str:
    """
    Trim by first comma OR first whitespace.
    Examples:
      '7.2.1, 7.2.2' -> '7.2.1'
      '7.2.1 build42' -> '7.2.1'
    """
    if not v:
        return ""
    v = v.strip()
    # cut at first comma
    v = v.split(",", 1)[0].strip()
    # cut at first whitespace
    v = re.split(r"\s+", v, 1)[0].strip()
    return v


def parse_iso_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def cf_value_to_str(v: Any) -> str:
    """Best-effort stringify for custom field values (enum/user/version/etc)."""
    if v is None:
        return ""
    # Value can be dict with name, fullName, localizedName, idReadable, etc.
    if isinstance(v, dict):
        for k in ("name", "localizedName", "fullName", "login", "idReadable"):
            if k in v and v[k]:
                return str(v[k])
        # Some multi-values come as {"values":[...]}
        if "values" in v and isinstance(v["values"], list):
            return ", ".join(cf_value_to_str(x) for x in v["values"])
        return str(v)
    if isinstance(v, list):
        return ", ".join(cf_value_to_str(x) for x in v)
    return str(v)


def get_custom_field(issue: Dict[str, Any], field_name: str) -> str:
    """Find custom field by name and return string value."""
    for cf in issue.get("customFields", []) or []:
        if cf.get("name") == field_name:
            return cf_value_to_str(cf.get("value"))
    return ""


def get_unique_filename(path: str) -> str:
    """
    If file exists, append (1), (2), (3) ... before extension,
    similar to standard OS behavior.
    """
    if not os.path.exists(path):
        return path

    directory = os.path.dirname(path)
    filename = os.path.basename(path)
    base, ext = os.path.splitext(filename)

    # Проверяем, есть ли уже "(n)" в конце имени
    match = re.match(r"^(.*)\((\d+)\)$", base.strip())
    if match:
        base_name = match.group(1).strip()
        counter = int(match.group(2)) + 1
    else:
        base_name = base
        counter = 1

    while True:
        new_filename = f"{base_name} ({counter}){ext}"
        new_path = os.path.join(directory, new_filename)
        if not os.path.exists(new_path):
            return new_path
        counter += 1


def fetch_issues(query: str, fields: str) -> List[Dict[str, Any]]:
    url = f"{BASE_URL.rstrip('/')}/api/issues"
    results: List[Dict[str, Any]] = []
    skip = 0
    page = 0

    while True:
        if MAX_PAGES is not None and page >= MAX_PAGES:
            print(f"Остановлено по лимиту страниц: {MAX_PAGES}")
            break

        print(f"Загрузка страницы {page + 1} (skip={skip}, top={PAGE_SIZE})")

        params = {
            "query": query,
            "fields": fields,
            "$top": PAGE_SIZE,
            "$skip": skip,
        }

        r = requests.get(url, headers=HEADERS, params=params, timeout=60)

        if r.status_code >= 400:
            print("HTTP", r.status_code)
            print(r.text)

        r.raise_for_status()

        batch = r.json()
        if not batch:
            print("Больше данных нет.")
            break

        results.extend(batch)

        if len(batch) < PAGE_SIZE:
            print("Последняя страница достигнута.")
            break

        skip += PAGE_SIZE
        page += 1

    print(f"Всего выгружено задач: {len(results)}")
    return results


ALLOWED_PROJECTS = {"VM", "BA", "DCI6"}

PROJECT_FILE_PREFIX = {
    "VM": "vm",
    "BA": "bill",
    "DCI6": "dci"
}

def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--project",
        required=False,
        help="Project short name (VM, BA, DCI6)"
    )
    args = parser.parse_args()

    if not args.project or args.project not in ALLOWED_PROJECTS:
        print("Неверно указан проект!")
        sys.exit(1)

    return args


# =======================
# Main
# =======================
def main():
    args = parse_args()
    project = args.project
    prefix = PROJECT_FILE_PREFIX[project]
    out_xlsx = f"{prefix}_defects_with_ps_links.xlsx"
    # --- Query for defects
    q_parts = [f"project: {project}", f"Type: {DEFECT_TYPE}"]
    if CREATED_FROM and not CREATED_TO:
        q_parts.append(f"Создана: {CREATED_FROM} .. *")

    elif CREATED_FROM and CREATED_TO:
        q_parts.append(f"Создана: {CREATED_FROM} .. {CREATED_TO}")

    elif CREATED_TO and not CREATED_FROM:
        q_parts.append(f"Создана: .. {CREATED_TO}")
    query = " ".join(q_parts)

    # --- Fields to fetch:
    # built-in: idReadable, summary, created, resolved
    # customFields: Статус, Приоритет, Релиз
    # links: only linked issues from PS are filtered on our side (simpler/safer),
    # but we fetch linked issues with their project shortName + PS version field.
    fields = (
        "id,idReadable,summary,created,resolved,"
        "customFields(name,value(name,localizedName,fullName,login,idReadable,values(name,localizedName,fullName,login,idReadable))),"
        "links(direction,linkType(name,localizedName),issues(idReadable,project(shortName),customFields(name,value(name,localizedName,fullName,login,idReadable,values(name,localizedName,fullName,login,idReadable)))))"
    )

    issues = fetch_issues(query=query, fields=fields)
    print(f"Fetched {project} defects: {len(issues)}")

    cutoff = parse_iso_date(RESOLVED_CUTOFF)

    total_fetched = len(issues)
    filtered_out = 0
    kept_total = 0
    kept_unresolved = 0

    rows = []

    for it in issues:
        vm_id = it.get("idReadable") or it.get("id")
        summary = it.get("summary", "")

        created = yt_dt(it.get("created"))
        resolved_str = yt_dt(it.get("resolved"))
        lifetime = calc_lifetime(created, resolved_str)
        quarter, month = calc_quarter_month(created)
        resolved_dt = parse_iso_date(resolved_str) if resolved_str else None

        # фильтрация
        if resolved_dt is not None and resolved_dt < cutoff:
            filtered_out += 1
            continue

        kept_total += 1
        if resolved_dt is None:
            kept_unresolved += 1

        status = get_custom_field(it, FIELD_STATUS)
        priority = get_custom_field(it, FIELD_PRIORITY)
        release = get_custom_field(it, FIELD_RELEASE)

        # Collect PS-linked issues + their version field
        ps_ids: List[str] = []
        ps_versions: List[str] = []

        for link in it.get("links", []) or []:
            for linked in link.get("issues", []) or []:
                prj = (linked.get("project") or {}).get("shortName")
                if prj != PS_PROJECT:
                    continue

                linked_id = linked.get("idReadable", "")
                version_raw = get_custom_field(linked, PS_VERSION_FIELD)
                version = normalize_ps_version(version_raw)

                if version:
                    ps_versions.append(version)

                ps_ids.append(linked_id)

                # “Relates to(OUTWARD): PS-123 [v1.2]”
                extra = f" [{version}]" if version else ""

        rows.append({
            "id": vm_id,
            "Summary": summary,
            "Status": status,
            "Priority": priority,
            "Created": created,
            "Quarter": quarter,
            "Month": month,
            "Resolved": resolved_str,
            "Lifetime": lifetime,
            "Release": release,
            "PS links (IDs)": ", ".join(ps_ids),
            f"PS_{PS_VERSION_FIELD}": ", ".join(ps_versions)
        })

    df = pd.DataFrame(rows)

    # Nice ordering
    if not df.empty:
        df = df.sort_values(["id"], ascending=[True])

    final_path = get_unique_filename(out_xlsx)
    df.to_excel(final_path, index=False)
    print(f"Saved: {final_path}")

    # --- Highlight PS_Версия when PS links exist but PS_Версия is empty
    wb = load_workbook(final_path)
    ws = wb.active  # ваш лист, если он один

    # Находим индексы колонок по заголовкам
    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    ps_links_col = None
    ps_ver_col = None

    # "PS links (IDs)" — как у тебя сейчас называется колонка
    for k in header.keys():
        if isinstance(k, str) and k.strip() == "PS links (IDs)":
            ps_links_col = header[k]
        if isinstance(k, str) and k.strip() == f"PS_{PS_VERSION_FIELD}":
            ps_ver_col = header[k]

    if ps_links_col is None or ps_ver_col is None:
        raise RuntimeError("Не найдены колонки 'PS links (IDs)' и/или 'PS_Версия' в итоговом файле.")

    coral_fill = PatternFill(fill_type="solid", start_color="FF7F50", end_color="FF7F50")  # Coral
    coral_count = 0

    for r in range(2, ws.max_row + 1):
        ps_links_val = ws.cell(row=r, column=ps_links_col).value
        ps_ver_cell = ws.cell(row=r, column=ps_ver_col)
        ps_ver_val = ps_ver_cell.value

        has_links = bool(str(ps_links_val).strip()) if ps_links_val is not None else False
        has_version = bool(str(ps_ver_val).strip()) if ps_ver_val is not None else False

        if has_links and not has_version:
            ps_ver_cell.fill = coral_fill
            coral_count += 1

    wb.save(final_path)

    print("\n===== SUMMARY =====")
    print(f"Всего получено из API:        {total_fetched}")
    print(f"Отфильтровано (resolved < {RESOLVED_CUTOFF}): {filtered_out}")
    print(f"Итого в файле:                {kept_total}")
    print(f"Из них unresolved:            {kept_unresolved}")
    print(f"Есть PS links, но нет версии): {coral_count}")
    print("===================")


if __name__ == "__main__":
    main()
